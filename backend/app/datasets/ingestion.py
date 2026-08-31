from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Literal

import pandas as pd
from openpyxl import load_workbook

from app.datasets.validation import ValidationError

NoticeCode = Literal[
    "encoding_detected",
    "mixed_date_formats",
    "ambiguous_date_values",
    "numeric_format_normalized",
    "multi_sheet_combined",
    "merged_header_flattened",
]


@dataclass(frozen=True)
class IngestionNotice:
    code: NoticeCode
    message: str
    column: str | None = None
    details: dict[str, object] = field(default_factory=dict)


@dataclass(frozen=True)
class IngestionResult:
    dataframe: pd.DataFrame
    notices: list[IngestionNotice]


_AMBIGUOUS_DATE = re.compile(r"^\s*(\d{1,2})[/-](\d{1,2})[/-](\d{2}|\d{4})(?:\s|$)")
_CURRENCY = re.compile(r"[$€£¥₹]")
_IDENTIFIER_NAME = re.compile(r"(^|_)(id|code|zip|postal|phone|account)(_|$)", re.IGNORECASE)


def parse_dataset(content: bytes, extension: str) -> IngestionResult:
    try:
        if extension == ".csv":
            frame, notices = _read_csv(content)
        elif extension == ".xlsx":
            frame, notices = _read_xlsx(content)
        else:
            raise ValidationError(f"Unsupported ingestion extension '{extension}'.")
    except ValidationError:
        raise
    except Exception as exc:  # noqa: BLE001 - converted into the upload API's stable 400 contract
        raise ValidationError(f"Could not parse file: {exc}") from exc

    frame, normalization_notices = normalize_dataframe(frame)
    return IngestionResult(frame, notices + normalization_notices)


def _read_csv(content: bytes) -> tuple[pd.DataFrame, list[IngestionNotice]]:
    encoding = _detect_encoding(content)
    text = content.decode(encoding)
    frame = pd.read_csv(io.StringIO(text))
    notices: list[IngestionNotice] = []
    if encoding not in {"utf-8", "utf-8-sig"}:
        notices.append(IngestionNotice(
            code="encoding_detected",
            message=f"CSV text was decoded as {encoding}; verify non-ASCII characters if this was unexpected.",
            details={"encoding": encoding},
        ))
    return frame, notices


def _detect_encoding(content: bytes) -> str:
    if content.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    try:
        content.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        pass
    for encoding in ("cp1252", "latin-1"):
        try:
            decoded = content.decode(encoding)
        except UnicodeDecodeError:
            continue
        if "\x00" not in decoded:
            return encoding
    raise ValidationError("Could not determine a safe CSV text encoding.")


def _read_xlsx(content: bytes) -> tuple[pd.DataFrame, list[IngestionNotice]]:
    workbook = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    frames: list[pd.DataFrame] = []
    notices: list[IngestionNotice] = []
    for sheet in workbook.worksheets:
        frame, flattened = _worksheet_to_frame(sheet)
        if frame.empty and len(frame.columns) == 0:
            continue
        frame = frame.dropna(how="all")
        if frame.empty:
            continue
        frame.attrs["source_sheet"] = sheet.title
        frames.append(frame)
        if flattened:
            notices.append(IngestionNotice(
                code="merged_header_flattened",
                message=f"Merged header cells on sheet '{sheet.title}' were flattened into column names.",
                details={"sheet": sheet.title},
            ))
    if not frames:
        raise ValidationError("Workbook contains no non-empty worksheets.")
    if len(frames) == 1:
        return frames[0], notices

    sheet_column = "__sheet__"
    while any(sheet_column in frame.columns for frame in frames):
        sheet_column = "_" + sheet_column
    combined: list[pd.DataFrame] = []
    sheet_names: list[str] = []
    for frame in frames:
        name = str(frame.attrs["source_sheet"])
        copy = frame.copy()
        copy.insert(0, sheet_column, name)
        combined.append(copy)
        sheet_names.append(name)
    notices.append(IngestionNotice(
        code="multi_sheet_combined",
        message=f"Combined {len(combined)} non-empty worksheets; source rows are identified by '{sheet_column}'.",
        details={"sheets": sheet_names, "sheet_column": sheet_column},
    ))
    return pd.concat(combined, ignore_index=True, sort=False), notices


def _worksheet_to_frame(sheet) -> tuple[pd.DataFrame, bool]:
    rows = list(sheet.iter_rows(values_only=True))
    while rows and all(value is None for value in rows[-1]):
        rows.pop()
    if not rows:
        return pd.DataFrame(), False

    width = max(len(row) for row in rows)
    padded = [list(row) + [None] * (width - len(row)) for row in rows]
    header_index = max(range(min(10, len(padded))), key=lambda index: _non_empty_count(padded[index]))
    flattened = False
    header = padded[header_index]
    if header_index > 0 and _has_merged_header(sheet, header_index):
        upper = pd.Series(padded[header_index - 1], dtype="object").ffill().tolist()
        header = [
            " | ".join(str(part).strip() for part in (upper[i], header[i]) if part is not None and str(part).strip())
            for i in range(width)
        ]
        flattened = True
    columns = _unique_columns(header)
    return pd.DataFrame(padded[header_index + 1 :], columns=columns), flattened


def _has_merged_header(sheet, zero_based_header_index: int) -> bool:
    upper_excel_row = zero_based_header_index
    return any(cell_range.min_row == upper_excel_row for cell_range in sheet.merged_cells.ranges)


def _non_empty_count(row: list[object]) -> int:
    return sum(value is not None and str(value).strip() != "" for value in row)


def _unique_columns(values: list[object]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for index, value in enumerate(values, start=1):
        base = str(value).strip() if value is not None and str(value).strip() else f"column_{index}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        result.append(base if count == 0 else f"{base}_{count + 1}")
    return result


def normalize_dataframe(frame: pd.DataFrame) -> tuple[pd.DataFrame, list[IngestionNotice]]:
    frame = frame.copy()
    notices: list[IngestionNotice] = []
    for column in frame.columns:
        if frame[column].dtype != object:
            continue
        numeric = _normalize_formatted_numeric(frame[column], str(column))
        if numeric is not None:
            frame[column] = numeric
            notices.append(IngestionNotice(
                code="numeric_format_normalized",
                column=str(column),
                message=f"Formatted numeric values in '{column}' were converted to numbers.",
            ))
            continue
        parsed_dates, date_notices = _normalize_dates(frame[column], str(column))
        if parsed_dates is not None:
            frame[column] = parsed_dates
            notices.extend(date_notices)
    return frame, notices


def _normalize_formatted_numeric(series: pd.Series, column: str) -> pd.Series | None:
    if _IDENTIFIER_NAME.search(column):
        return None
    non_null = series.dropna()
    if non_null.empty:
        return None
    text = non_null.astype(str).str.strip()
    formatted = text.str.contains(_CURRENCY).any() or text.str.contains("%", regex=False).any()
    formatted = formatted or text.str.match(r"^\(?[-+]?\d{1,3}([,.]\d{3})+([,.]\d+)?\)?$").any()
    if not formatted:
        return None
    parsed = text.map(_parse_formatted_number)
    if parsed.notna().mean() < 0.9:
        return None
    result = pd.Series(pd.NA, index=series.index, dtype="Float64")
    result.loc[non_null.index] = parsed.astype("Float64")
    return result


def _parse_formatted_number(value: str) -> float | None:
    text = value.strip().replace("\u00a0", "").replace(" ", "")
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    is_percent = text.endswith("%")
    text = text.rstrip("%")
    text = _CURRENCY.sub("", text)
    if "," in text and "." in text:
        decimal = "," if text.rfind(",") > text.rfind(".") else "."
        thousands = "." if decimal == "," else ","
        text = text.replace(thousands, "").replace(decimal, ".")
    elif "," in text:
        parts = text.split(",")
        text = "".join(parts) if len(parts[-1]) == 3 else ".".join(parts)
    elif text.count(".") > 1:
        parts = text.split(".")
        text = "".join(parts) if all(len(part) == 3 for part in parts[1:]) else text
    try:
        number = float(text)
    except ValueError:
        return None
    if negative:
        number = -number
    return number / 100 if is_percent else number


def _normalize_dates(series: pd.Series, column: str) -> tuple[pd.Series | None, list[IngestionNotice]]:
    non_null = series.dropna()
    if non_null.empty:
        return None, []
    text = non_null.astype(str).str.strip()
    parsed = pd.to_datetime(text, errors="coerce", format="mixed", dayfirst=False)
    if parsed.notna().mean() < 0.9:
        return None, []

    formats = {_date_shape(value) for value in text if _date_shape(value)}
    ambiguous = [value for value in text if _is_ambiguous_date(value)]
    result = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
    result.loc[non_null.index] = parsed
    notices: list[IngestionNotice] = []
    if len(formats) > 1:
        notices.append(IngestionNotice(
            code="mixed_date_formats",
            column=column,
            message=f"Column '{column}' contains multiple date formats parsed with one column-level policy.",
            details={"formats": sorted(formats)},
        ))
    if ambiguous:
        notices.append(IngestionNotice(
            code="ambiguous_date_values",
            column=column,
            message=f"Column '{column}' contains {len(ambiguous)} ambiguous day/month value(s); month-first interpretation was used.",
            details={"ambiguous_count": len(ambiguous), "policy": "month_first"},
        ))
    return result, notices


def _date_shape(value: str) -> str | None:
    value = value.strip()
    if re.match(r"^\d{4}-\d{1,2}-\d{1,2}", value):
        return "year-first"
    if re.match(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", value):
        return "numeric-day-month"
    if re.search(r"[A-Za-z]{3,}", value):
        return "named-month"
    return None


def _is_ambiguous_date(value: str) -> bool:
    match = _AMBIGUOUS_DATE.match(value)
    if not match:
        return False
    first, second = int(match.group(1)), int(match.group(2))
    return first != second and 1 <= first <= 12 and 1 <= second <= 12
