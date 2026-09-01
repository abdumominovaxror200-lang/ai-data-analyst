from __future__ import annotations

import io
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.data_quality_gate import evaluate_data_quality
from app.datasets.storage import DatasetRecord
from app.tools.report import generate_report

_NAVY = "172554"
_BLUE = "2563EB"
_LIGHT_GRAY = "F1F5F9"
_WHITE = "FFFFFF"
_RED = "FEE2E2"
_AMBER = "FEF3C7"
_GREEN = "DCFCE7"
_THIN_GRAY = Side(style="thin", color="CBD5E1")
_MAX_CELL_TEXT = 32_000


def build_excel_report(record: DatasetRecord) -> bytes:
    """Build a deterministic, styled workbook without running any provider."""
    report = generate_report(record.df, record.id, record.original_filename)
    caveats, quality_limitations = evaluate_data_quality(record.df)
    analysis = record.latest_analysis

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    caveat_sheet = workbook.create_sheet("Caveats")
    fact_sheet = workbook.create_sheet("Fact ledger")
    chart_sheet = workbook.create_sheet("Charts")
    limitation_sheet = workbook.create_sheet("Limitations")

    _build_summary(summary, report, analysis)
    _build_caveats(caveat_sheet, caveats, record.ingestion_notices or [])
    _build_fact_ledger(fact_sheet, analysis)
    _build_charts(chart_sheet, report)
    _build_limitations(limitation_sheet, quality_limitations, analysis)
    for sheet in workbook.worksheets:
        _finish_sheet(sheet)

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def excel_report_filename(record: DatasetRecord) -> str:
    stem = Path(record.original_filename).stem or "dataset"
    safe = "".join(
        character if character.isascii() and (character.isalnum() or character in "-_") else "_"
        for character in stem
    )
    return f"{safe[:100] or 'dataset'}-analysis-report.xlsx"


def _build_summary(sheet, report: dict[str, Any], analysis: Any | None) -> None:
    _title(sheet, "AI Data Analyst — Excel Report", "Summary")
    rows = [
        ("Dataset", report["filename"]),
        ("Dataset ID", report["dataset_id"]),
        ("Generated at", report["generated_at"]),
        ("Rows", report["overview"]["rows"]),
        ("Columns", report["overview"]["columns"]),
        ("Missing values", report["overview"]["missing_total"]),
        ("Duplicate rows", report["overview"]["duplicate_rows"]),
    ]
    _write_table(sheet, 4, ["Metric", "Value"], rows)
    row = 13
    sheet.cell(row, 1, "Key findings").font = Font(bold=True, size=12, color=_NAVY)
    row += 1
    for finding in report["key_findings"]:
        sheet.cell(row, 1, _safe_cell(finding))
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        sheet.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    row += 1
    if analysis is not None:
        sheet.cell(row, 1, "Latest grounded analysis").font = Font(bold=True, size=12, color=_NAVY)
        row += 1
        sheet.cell(row, 1, _safe_cell(analysis.final_answer_text))
        sheet.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=6)
        sheet.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row].height = 80
    else:
        sheet.cell(row, 1, "No Deep Reasoning analysis has been run for this dataset in the current server session.")
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)


def _build_caveats(sheet, caveats, ingestion_notices: list[Any]) -> None:
    _title(sheet, "Data Caveats", "Caveats")
    rows: list[tuple[Any, ...]] = [
        ("Duplicates", "duplicate rows", caveats.duplicate_row_count, f"{caveats.duplicate_pct}%"),
        ("Rows", "rows dropped", caveats.rows_dropped, caveats.rows_dropped_note),
    ]
    rows.extend(
        ("Coverage", item.column, item.non_null_rows, f"{item.coverage_pct}% of {item.total_rows} rows")
        for item in caveats.column_coverage
    )
    rows.extend(("Type anomaly", "dataset", "", item) for item in caveats.type_anomalies)
    rows.extend(
        ("Ingestion", notice.column or "dataset", notice.code, notice.message)
        for notice in ingestion_notices
    )
    _write_table(sheet, 4, ["Category", "Field", "Value", "Details"], rows)


def _build_fact_ledger(sheet, analysis: Any | None) -> None:
    _title(sheet, "Fact Ledger", "Fact ledger")
    headers = ["Fact ID", "Value", "Unit", "Tool", "Result path", "Row count", "Parameters", "Filters", "Source hash", "Timestamp"]
    facts = list(analysis.fact_ledger.facts) if analysis is not None and analysis.fact_ledger else []
    rows = [
        (
            fact.id, fact.value, fact.unit or "", fact.tool, fact.result_path,
            fact.row_count if fact.row_count is not None else "",
            _json(fact.params), _json(fact.filters), fact.source_hash, fact.ts.isoformat(),
        )
        for fact in facts
    ]
    if not rows:
        rows = [("No facts available", "", "", "", "", "", "", "", "", "")]
    _write_table(sheet, 4, headers, rows)
    for cell in sheet["B"][4:]:
        cell.number_format = "#,##0.00########"


def _build_charts(sheet, report: dict[str, Any]) -> None:
    _title(sheet, "Charts", "Charts")
    anomalies = report.get("anomalies", [])[:20]
    correlations = report.get("correlations", [])[:20]
    row = 4
    if anomalies:
        rows = [(item.get("column"), item.get("anomaly_count"), item.get("anomaly_pct")) for item in anomalies]
        _write_table(sheet, row, ["Anomaly column", "Count", "Percent"], rows)
        chart = BarChart()
        chart.title = "Anomalies by column"
        chart.y_axis.title = "Count"
        chart.add_data(Reference(sheet, min_col=2, min_row=row, max_row=row + len(rows)), titles_from_data=True)
        chart.set_categories(Reference(sheet, min_col=1, min_row=row + 1, max_row=row + len(rows)))
        chart.height, chart.width = 7, 13
        sheet.add_chart(chart, "E4")
        row += len(rows) + 3
    if correlations:
        rows = [
            (f"{item.get('column_a')} ↔ {item.get('column_b')}", item.get("correlation"), abs(item.get("correlation", 0)))
            for item in correlations
        ]
        _write_table(sheet, row, ["Correlation pair", "r", "Absolute r"], rows)
        chart = BarChart()
        chart.title = "Strongest correlations"
        chart.x_axis.title = "Absolute correlation"
        chart.add_data(Reference(sheet, min_col=3, min_row=row, max_row=row + len(rows)), titles_from_data=True)
        chart.set_categories(Reference(sheet, min_col=1, min_row=row + 1, max_row=row + len(rows)))
        chart.height, chart.width = 7, 13
        sheet.add_chart(chart, f"E{row}")
    if not anomalies and not correlations:
        sheet.cell(4, 1, "No chartable anomaly or correlation results are available.")


def _build_limitations(sheet, quality_limitations: list[Any], analysis: Any | None) -> None:
    _title(sheet, "Limitations", "Limitations")
    limitations = list(quality_limitations)
    if analysis is not None:
        limitations.extend(analysis.limitations)
    unique: dict[tuple[str, str, str], Any] = {}
    for limitation in limitations:
        unique[(limitation.category, limitation.severity, limitation.text)] = limitation
    rows = [
        (item.severity, item.category, item.text, ", ".join(item.affected_findings))
        for item in unique.values()
    ]
    if not rows:
        rows = [("none", "none", "No limitations were identified for this report.", "")]
    _write_table(sheet, 4, ["Severity", "Category", "Limitation", "Affected findings"], rows)
    for row_number in range(5, 5 + len(rows)):
        severity = sheet.cell(row_number, 1).value
        color = _RED if severity == "blocks_conclusion" else _AMBER if severity == "reduces_confidence" else _GREEN
        for cell in sheet[row_number]:
            cell.fill = PatternFill("solid", fgColor=color)


def _title(sheet, text: str, section: str) -> None:
    sheet.sheet_properties.tabColor = _BLUE
    sheet.merge_cells("A1:F1")
    cell = sheet["A1"]
    cell.value = text
    cell.font = Font(size=18, bold=True, color=_WHITE)
    cell.fill = PatternFill("solid", fgColor=_NAVY)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet["A2"] = f"Section: {section}"
    sheet["A2"].font = Font(italic=True, color="64748B")


def _write_table(sheet, start_row: int, headers: list[str], rows: Iterable[tuple[Any, ...]]) -> None:
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(start_row, column, header)
        cell.font = Font(bold=True, color=_WHITE)
        cell.fill = PatternFill("solid", fgColor=_BLUE)
        cell.alignment = Alignment(wrap_text=True)
        cell.border = Border(bottom=_THIN_GRAY)
    last_row = start_row
    for row_number, values in enumerate(rows, start=start_row + 1):
        last_row = row_number
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column, _safe_cell(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=_LIGHT_GRAY if row_number % 2 == 0 else _WHITE)
            cell.border = Border(bottom=_THIN_GRAY)
    sheet.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{max(last_row, start_row)}"
    sheet.freeze_panes = f"A{start_row + 1}"


def _finish_sheet(sheet) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        width = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
        sheet.column_dimensions[column_letter].width = min(max(width + 2, 12), 50)
    sheet.sheet_view.showGridLines = False


def _safe_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    text = str(value)[:_MAX_CELL_TEXT]
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def _json(value: Any) -> str:
    return json.dumps(value, sort_keys=True, ensure_ascii=False, default=str)
