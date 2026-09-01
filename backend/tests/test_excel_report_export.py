from __future__ import annotations

import io
from datetime import datetime, timezone
from types import SimpleNamespace

import pandas as pd
from openpyxl import load_workbook

from app.datasets.ingestion import IngestionNotice
from app.datasets.storage import DatasetRecord
from app.reasoning.contracts import Fact, FactLedger, Limitation
from app.reports.excel_export import build_excel_report, excel_report_filename
from tests.conftest import make_csv_bytes


def _record() -> DatasetRecord:
    values = list(range(1, 31)) + [500]
    frame = pd.DataFrame({
        "=unsafe_metric": values,
        "cost": [value * 2 for value in values],
        "region": ["North"] * 15 + ["South"] * 16,
    })
    fact = Fact(
        id="fact_abc",
        value=500,
        unit=None,
        tool="describe_data",
        params={"columns": ["=unsafe_metric"]},
        filters=[],
        row_count=len(frame),
        source_hash="a" * 64,
        ts=datetime(2026, 1, 1, tzinfo=timezone.utc),
        result_path="result.columns.=unsafe_metric.max",
    )
    record = DatasetRecord(
        id="dataset-1",
        original_filename="sales input.csv",
        extension=".csv",
        uploaded_at=datetime.now(timezone.utc),
        df=frame,
        stored_path="unused",
        ingestion_notices=[IngestionNotice(
            code="encoding_detected",
            message="CSV text was decoded as cp1252.",
            details={"encoding": "cp1252"},
        )],
    )
    record.latest_analysis = SimpleNamespace(
        final_answer_text="Verified analysis summary.",
        fact_ledger=FactLedger(facts=[fact]),
        limitations=[Limitation(
            category="methodological",
            severity="reduces_confidence",
            text="One methodological limitation remains.",
        )],
    )
    return record


def test_formatted_workbook_contains_all_required_sheets_and_content():
    workbook = load_workbook(io.BytesIO(build_excel_report(_record())))
    assert workbook.sheetnames == ["Summary", "Caveats", "Fact ledger", "Charts", "Limitations"]
    assert workbook["Summary"]["A1"].value == "AI Data Analyst — Excel Report"
    assert workbook["Summary"]["A1"].font.bold is True
    assert workbook["Summary"]["A1"].fill.fgColor.rgb.endswith("172554")
    assert "Verified analysis summary." in [cell.value for row in workbook["Summary"] for cell in row]
    assert "fact_abc" in [cell.value for row in workbook["Fact ledger"] for cell in row]
    assert "One methodological limitation remains." in [cell.value for row in workbook["Limitations"] for cell in row]
    assert "encoding_detected" in [cell.value for row in workbook["Caveats"] for cell in row]


def test_workbook_has_filters_freeze_panes_native_charts_and_safe_cells():
    workbook = load_workbook(io.BytesIO(build_excel_report(_record())))
    for name in ("Summary", "Caveats", "Fact ledger", "Limitations"):
        assert workbook[name].freeze_panes == "A5"
        assert workbook[name].auto_filter.ref
        assert workbook[name].sheet_view.showGridLines is False
    chart_sheet = workbook["Charts"]
    assert chart_sheet._charts  # native Excel charts, not screenshots
    values = [cell.value for row in chart_sheet for cell in row if isinstance(cell.value, str)]
    assert any(value.startswith("'=unsafe_metric") for value in values)
    assert not any(value.startswith("=unsafe_metric") for value in values)


def test_export_without_reasoning_still_has_explicit_empty_fact_and_limitation_sections():
    record = _record()
    record.latest_analysis = None
    workbook = load_workbook(io.BytesIO(build_excel_report(record)))
    summary_values = [cell.value for row in workbook["Summary"] for cell in row]
    fact_values = [cell.value for row in workbook["Fact ledger"] for cell in row]
    assert any("No Deep Reasoning analysis" in str(value) for value in summary_values)
    assert "No facts available" in fact_values


def test_export_filename_is_safe_and_bounded():
    record = _record()
    record.original_filename = "../Q4 sales?.csv"
    filename = excel_report_filename(record)
    assert filename == "Q4_sales_-analysis-report.xlsx"
    assert "/" not in filename and "\\" not in filename
    record.original_filename = "mijozlar_ҳисоботи.xlsx"
    assert excel_report_filename(record).isascii()


def test_excel_report_api_returns_downloadable_workbook(client, sample_df):
    uploaded = client.post(
        "/api/datasets/upload",
        files={"file": ("sales.csv", make_csv_bytes(sample_df), "text/csv")},
    ).json()
    response = client.get(f"/api/reports/{uploaded['dataset_id']}/excel")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "sales-analysis-report.xlsx" in response.headers["content-disposition"]
    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == ["Summary", "Caveats", "Fact ledger", "Charts", "Limitations"]


def test_unknown_dataset_excel_export_returns_404(client):
    response = client.get("/api/reports/not-a-real-dataset/excel")
    assert response.status_code == 404
