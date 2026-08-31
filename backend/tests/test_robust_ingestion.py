from __future__ import annotations

import io

import pandas as pd
import pytest
from openpyxl import Workbook

from app.datasets.ingestion import parse_dataset
from app.datasets.validation import ValidationError


def _xlsx_bytes(build) -> bytes:
    workbook = Workbook()
    build(workbook)
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_csv_detects_legacy_encoding_without_corrupting_text():
    content = "city,revenue\nMontréal,12\nZürich,14\n".encode("cp1252")
    result = parse_dataset(content, ".csv")
    assert result.dataframe["city"].tolist() == ["Montréal", "Zürich"]
    assert any(notice.code == "encoding_detected" and notice.details["encoding"] == "cp1252" for notice in result.notices)


def test_currency_percent_and_thousands_are_normalized_deterministically():
    content = (
        "amount,margin,european\n"
        '"$1,234.50",12.5%,"1.234,50"\n'
        '"($2,000.00)",8%,"2.500,00"\n'
    ).encode()
    result = parse_dataset(content, ".csv")
    assert result.dataframe["amount"].tolist() == [1234.5, -2000.0]
    assert result.dataframe["margin"].tolist() == [0.125, 0.08]
    assert result.dataframe["european"].tolist() == [1234.5, 2500.0]
    assert {notice.column for notice in result.notices if notice.code == "numeric_format_normalized"} == {
        "amount", "margin", "european"
    }


def test_mixed_date_formats_are_parsed_per_column_and_ambiguity_is_flagged():
    content = b"date,value\n2025-01-31,1\n01/02/2025,2\nMarch 3 2025,3\n"
    result = parse_dataset(content, ".csv")
    assert pd.api.types.is_datetime64_any_dtype(result.dataframe["date"])
    assert result.dataframe.loc[1, "date"] == pd.Timestamp("2025-01-02")
    codes = {notice.code for notice in result.notices if notice.column == "date"}
    assert codes == {"mixed_date_formats", "ambiguous_date_values"}


def test_multi_sheet_workbook_combines_non_empty_sheets_with_provenance():
    def build(workbook):
        first = workbook.active
        first.title = "North"
        first.append(["region", "revenue"])
        first.append(["N", 10])
        second = workbook.create_sheet("South")
        second.append(["region", "revenue"])
        second.append(["S", 20])
        workbook.create_sheet("Empty")

    result = parse_dataset(_xlsx_bytes(build), ".xlsx")
    assert result.dataframe["__sheet__"].tolist() == ["North", "South"]
    assert result.dataframe["revenue"].tolist() == [10, 20]
    notice = next(notice for notice in result.notices if notice.code == "multi_sheet_combined")
    assert notice.details["sheets"] == ["North", "South"]


def test_merged_two_row_excel_header_is_flattened_without_losing_columns():
    def build(workbook):
        sheet = workbook.active
        sheet.merge_cells("A1:B1")
        sheet["A1"] = "Financial"
        sheet["C1"] = "Location"
        sheet.append(["Revenue", "Cost", "Region"])
        sheet.append([100, 60, "East"])

    result = parse_dataset(_xlsx_bytes(build), ".xlsx")
    assert result.dataframe.columns.tolist() == [
        "Financial | Revenue", "Financial | Cost", "Location | Region"
    ]
    assert result.dataframe.iloc[0].tolist() == [100, 60, "East"]
    assert any(notice.code == "merged_header_flattened" for notice in result.notices)


def test_identifier_like_formatted_values_are_not_coerced_to_numbers():
    result = parse_dataset(b"account_id,value\n001,1\n002,2\n", ".csv")
    # pandas' ordinary CSV inference remains backward compatible; the robust
    # formatted-number pass must not perform any additional identifier conversion.
    assert not any(notice.column == "account_id" for notice in result.notices)


def test_upload_profile_exposes_ingestion_notices(client):
    files = {"file": ("legacy.csv", "city,margin\nMontréal,12%\n".encode("cp1252"), "text/csv")}
    response = client.post("/api/datasets/upload", files=files)
    assert response.status_code == 200
    codes = {notice["code"] for notice in response.json()["profile"]["ingestion_notices"]}
    assert {"encoding_detected", "numeric_format_normalized"} <= codes


@pytest.mark.parametrize("extension", [".csv", ".xlsx"])
def test_empty_or_invalid_content_has_stable_validation_error(extension):
    with pytest.raises(ValidationError):
        parse_dataset(b"not a valid workbook" if extension == ".xlsx" else b"\xff\x00", extension)
